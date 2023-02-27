import os
import sys

import openpyxl
from PyQt5 import QtWidgets
from PyQt5.QtCore import QTimer, QEventLoop, Qt
from PyQt5.QtGui import QPixmap, QMovie
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, \
    QFileDialog, QLabel, QMessageBox, QInputDialog, QSizePolicy, QTableWidgetItem, QAbstractItemView

from JG_design import Ui_JEOPARDYGAMEMAKER
from database import DataBase
from layout_creator import LayoutCreator
from player import Player
from timer import Timer
from xlsx_handler import Handler


class MyProgram(QMainWindow, Ui_JEOPARDYGAMEMAKER):
    def __init__(self):
        super().__init__()
        self.path = None
        self.xlsx_path = None
        self.setupUi(self)
        self.number_of_players = 0
        self.list_of_lbl_with_players_in_game = []
        self.list_of_widget_with_players_in_answer = []
        self.list_of_lbl_themes = []
        self.list_of_players = []
        self.questions_pack = []
        self.list_btn_true = []
        self.list_btn_false = []
        self.list_btn_ans = []
        self.list_players_lbl = []
        self.list_of_lbl_with_players_name = []
        self.btn_back_in_standings_to_initial_window.clicked.connect(self.go_to_initial_window)
        self.btn_next_in_game.clicked.connect(self.next_in_game)
        self.btn_start.clicked.connect(self.go_start_game)
        self.btn_start_game.clicked.connect(self.go_game)
        self.btn_input_file.clicked.connect(self.input_file)
        self.btn_add_players.clicked.connect(self.add_player)
        self.btn_create_game.clicked.connect(self.create_game)
        self.btn_save_question.clicked.connect(self.push_btn)
        self.btn_back_create_game.clicked.connect(self.go_start_game)
        self.btn_del_player.clicked.connect(self.del_player)
        self.btn_standings.clicked.connect(self.go_standings)
        self.btn_back_to_initial_window.clicked.connect(self.go_to_initial_window)
        self.flag = False
        self.timer = None
        self.cost = None
        self.can_start = False
        self.current_question = None
        self.lineEdit_player_name.returnPressed.connect(self.btn_add_players.click)
        self.lineedit_question.returnPressed.connect(self.btn_save_question.click)
        self.desktop = QApplication.desktop()
        self.screenRect = self.desktop.screenGeometry()
        self.height_window = self.screenRect.height()
        self.width_window = self.screenRect.width()
        self.database = DataBase("application files/users_db.sqlite")
        self.text_in_question_label = True
        self.count_of_questions_used = 0
        self.btn_end_game.clicked.connect(self.go_to_initial_window)

    def create_question(self, row, column, btn):
        self.current_question = self.questions_pack[row][column - 1]
        self.path = os.path.join(os.path.dirname(self.xlsx_path), 'files', self.current_question[0])
        if os.path.exists(self.path):
            self.set_img(self.path)
            self.text_in_question_label = False
        else:
            self.lbl_question.setText(self.current_question[0])
            self.text_in_question_label = True
        self.stack.setCurrentIndex(5)
        self.cost = column * 100
        btn.setEnabled(False)
        self.timer.flag = True

    def change_progress(self, value):
        self.progress.setValue(value)

    def make_buttons(self):
        for row in range(len(self.questions_pack)):
            cost = 0
            for column in range(1, len(self.questions_pack[row]) + 1):
                btn = QPushButton(self.questions_layout)
                size_policy_btn = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding,
                                                        QtWidgets.QSizePolicy.Expanding)
                btn.setSizePolicy(size_policy_btn)
                cost += 100
                btn.setText(str(cost))
                self.gridLayout_3.addWidget(btn, row, column, 1, 1)
                btn.clicked.connect(
                    lambda pass_value, row_value=row, column_value=column,
                           button=btn: self.create_question(row_value, column_value, button))

    def make_players_in_game(self):
        layout_creator = LayoutCreator(self.checkbox_input_ans.isChecked())

        for player in self.list_of_players:
            lbl_first = QLabel(self.players_layout)
            widget = QtWidgets.QWidget(self.players_in_game)
            widget.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Preferred)

            if not self.checkbox_input_ans.isChecked():
                btn_true, btn_false, lbl_first = \
                    layout_creator.create_layout_with_players(player, widget, lbl_first)
                self.layout_with_all_players.addWidget(lbl_first)
                self.list_btn_true.append(btn_true)
                self.list_btn_false.append(btn_false)
                self.layout_with_all_players_to_answer.addWidget(widget)
                btn_true.clicked.connect(
                    lambda pass_value, player_value=player, label=lbl_first: self.true_answer(
                        player_value, label))
                btn_false.clicked.connect(
                    lambda pass_value, player_value=player, label=lbl_first, button_false=btn_false,
                           button_true=btn_true: self.false_answer(player_value, label,
                                                                   button_false, button_true))
            else:
                btn_ans, lbl_first = \
                    layout_creator.create_layout_with_players(player, widget, lbl_first)
                self.layout_with_all_players.addWidget(lbl_first)
                self.list_btn_ans.append(btn_ans)
                self.layout_with_all_players_to_answer.addWidget(widget)
                btn_ans.clicked.connect(
                    lambda pass_value, player_value=player, label=lbl_first,
                           button_false=btn_ans: self.input_answer(
                        player_value, label, button_false))
            self.list_of_lbl_with_players_in_game.append(lbl_first)
            self.list_of_widget_with_players_in_answer.append(widget)

    def input_answer(self, player, lbl_first, btn_false):
        self.timer.flag = False
        text, ok = QInputDialog.getText(self, 'Ввод ответа', player.name + ', введите ответ:')
        if ok:
            if text.lower().strip() == self.current_question[1].lower().strip():
                self.true_answer(player, lbl_first)
            else:
                self.false_answer(player, lbl_first, btn_false, True)
                self.timer.flag = True
        else:
            self.false_answer(player, lbl_first, btn_false, True)
            self.timer.flag = True

    def set_img(self, path):
        extension = self.current_question[0].split('.')[-1]
        if extension.lower() in ['jpg', 'png', 'jpeg', 'gif', 'mp4']:
            if extension.lower() == 'gif':
                img = QMovie(path)
                self.lbl_question.setMovie(img)
                img.start()
            else:
                img = QPixmap(path)
                img = img.scaled(int(0.8 * self.width_window), int(0.8 * self.height_window),
                                 Qt.KeepAspectRatio)
                self.lbl_question.setPixmap(img)

    def next_in_game(self):
        for btn in self.list_btn_true:
            btn.setEnabled(False)
            btn.setStyleSheet('''background-color: rgb(255, 74, 42)''')
        for btn in self.list_btn_false:
            btn.setEnabled(False)
            btn.setStyleSheet('''background-color: rgb(255, 74, 42)''')
        for btn in self.list_btn_ans:
            btn.setEnabled(False)
            btn.setStyleSheet('''background-color: rgb(255, 74, 42)''')

        self.timer.flag = False
        if self.lbl_question.text() == self.current_question[0] or not self.text_in_question_label:
            path = os.path.join(os.path.dirname(self.xlsx_path), 'files',
                                str(self.current_question[1]))
            if os.path.exists(path):
                self.set_img(path)
            else:
                self.lbl_question.setText('Правильный ответ:\n\n' + str(self.current_question[1]))
            self.text_in_question_label = True
        else:
            self.timer.count = 0
            self.stack.setCurrentIndex(3)
            self.change_count_played_questions()

            for btn in self.list_btn_true:
                btn.setEnabled(True)
                btn.setStyleSheet('''background-color: rgb(66, 239, 214)''')
            for btn in self.list_btn_false:
                btn.setEnabled(True)
                btn.setStyleSheet('''background-color: rgb(66, 239, 214)''')
            for btn in self.list_btn_ans:
                btn.setEnabled(True)
                btn.setStyleSheet('''background-color: rgb(66, 239, 214)''')

    def go_to_initial_window(self):
        self.stack.setCurrentIndex(0)

    def true_answer(self, player, lbl_first):
        player.plus_score(self.cost)
        lbl_first.setText(str(player.name + '\nСчёт: ' + '\n' + str(player.get_score())))
        self.next_in_game()

    def false_answer(self, player, lbl_first, btn_false, btn_true):
        player.minus_score(self.cost)
        if not self.checkbox_input_ans.isChecked():
            btn_false.setEnabled(False)
            btn_true.setEnabled(False)
            btn_false.setStyleSheet('''background-color: rgb(255, 74, 42)''')
            btn_true.setStyleSheet('''background-color: rgb(255, 74, 42)''')
        else:
            btn_false.setEnabled(False)
            btn_false.setStyleSheet('''background-color: rgb(255, 74, 42)''')
        lbl_first.setText(str(player.name + '\nСчёт: ' + '\n' + str(player.get_score())))

    def input_file(self):
        for lbl in self.list_of_lbl_themes:
            lbl.clear()
        for index in reversed(range(self.themes_with_theme_names.count())):
            self.themes_with_theme_names.itemAt(index).widget().setParent(None)
        self.xlsx_path = QFileDialog.getOpenFileName(self, 'Выбрать файл', '', '*.xlsx')[0]
        handler = Handler(self.xlsx_path)
        function_response = handler.processing()
        if function_response != "Error":
            themes, self.questions_pack = function_response
            self.btn_input_file.setText(self.xlsx_path.split("/")[-1])
            for index in range(len(themes)):
                lbl = QLabel(self.themes_layout)
                lbl.setText(themes[index])
                lbl.setStyleSheet(
                    '''background-color: rgb(171, 223, 255); border: None; border-radius: 10px;''')
                lbl.setAlignment(Qt.AlignCenter)
                self.list_of_lbl_themes.append(lbl)
                self.themes_with_theme_names.addWidget(lbl)

            self.make_buttons()

            if self.can_start:
                self.btn_start_game.setEnabled(True)
            else:
                self.can_start = True
        else:
            QMessageBox.critical(self, "Ошибка ", 'Ошибка импорта файла! Попробуйте ещё раз!', QMessageBox.Ok)

    def go_start_game(self):
        self.stack.setCurrentIndex(2)

    def push_btn(self):
        self.flag = True

    def create_game(self):
        def wait_push_btn():
            while not self.flag:
                loop_timer = QEventLoop()
                QTimer.singleShot(1, loop_timer.quit)
                loop_timer.exec_()

        self.label_with_number_question.setText('Введите название пакета тем')
        self.stack.setCurrentIndex(6)
        wait_push_btn()
        self.flag = False
        name = self.lineedit_question.text()
        self.lineedit_question.clear()

        file = openpyxl.Workbook()
        file.create_sheet(title='Вопросы', index=0)
        file.create_sheet(title='Ответы', index=1)
        sheet1 = file['Вопросы']
        sheet2 = file['Ответы']

        for index in range(66, 71):
            sheet1[chr(index) + '1'] = (index - 65) * 100
            sheet2[chr(index) + '1'] = (index - 65) * 100

        themes = []

        for index in range(2, 8):
            self.flag = False
            self.label_with_number_question.setText(f'Ведите тему номер {index - 1}')
            while not self.flag:
                loop = QEventLoop()
                QTimer.singleShot(1, loop.quit)
                loop.exec_()
            sheet1['A' + str(index)] = self.lineedit_question.text()
            sheet2['A' + str(index)] = self.lineedit_question.text()
            themes.append(self.lineedit_question.text())
            self.lineedit_question.clear()

        words = {'B': 100, 'C': 200, 'D': 300, 'E': 400, 'F': 500}

        for index in range(2, 8):
            for j in words:
                self.flag = False
                self.label_with_number_question.setText(
                    f'Ведите вопрос на тему "{themes[index - 2]}" за {words[j]}')
                wait_push_btn()
                sheet1[j + str(index)] = self.lineedit_question.text()
                self.lineedit_question.clear()
                self.label_with_number_question.setText(
                    f'Ведите ответ на тему "{themes[index - 2]}" за {words[j]}')
                wait_push_btn()
                sheet2[j + str(index)] = self.lineedit_question.text()
                self.lineedit_question.clear()

        file.save(f'{name}.xlsx')
        self.stack.setCurrentIndex(2)

    def go_create_game(self):
        self.stack.setCurrentIndex(6)

    def create_table(self):
        table_header = ['Имя', 'Количество игр', 'Количество баллов', 'Количество побед']
        self.table.setColumnCount(len(table_header))
        self.table.setHorizontalHeaderLabels(table_header)
        self.table.setRowCount(0)
        data_from_database = self.database.get_data()
        if data_from_database != "Error":
            table_elementsreferences = reversed(
                sorted(data_from_database, key=lambda player_count: player_count[2]))
            for row_value, row in enumerate(table_elementsreferences):
                self.table.setRowCount(self.table.rowCount() + 1)
                for column_value, element in enumerate(row):
                    self.table.setItem(row_value, column_value, QTableWidgetItem(str(element)))
            self.table.resizeColumnsToContents()
            self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        else:
            QMessageBox.critical(self, "Ошибка ", 'Ошибка базы данных! Проверьте, есть ли рядо с вашим файлом папка'
                                                  ' "application files" с файлом "users_db.sqlite"!', QMessageBox.Ok)
            self.stack.setCurrentIndex(0)

    def go_standings(self):
        self.stack.setCurrentIndex(7)
        self.create_table()

    def clear_gamemaker_data(self):
        for wgt in self.list_of_widget_with_players_in_answer:
            wgt.deleteLater()
        for lbl in self.list_of_lbl_with_players_in_game:
            lbl.deleteLater()
        for lbl in self.list_of_lbl_with_players_name:
            lbl.deleteLater()
        self.btn_input_file.setText("Выбрать файл с игрой")
        self.btn_start_game.setEnabled(False)
        self.btn_del_player.setEnabled(False)
        self.path = None
        self.xlsx_path = None
        self.number_of_players = 0
        self.list_of_lbl_themes = []
        self.list_of_lbl_with_players_in_game = []
        self.list_of_widget_with_players_in_answer = []
        self.list_of_lbl_with_players_name = []
        self.list_of_players = []
        self.questions_pack = []
        self.list_btn_true = []
        self.list_btn_false = []
        self.list_btn_ans = []
        self.list_players_lbl = []
        self.flag = False
        del self.timer
        self.timer = None
        self.cost = None
        self.can_start = False
        self.current_question = None
        self.text_in_question_label = True
        self.count_of_questions_used = 0

    def change_count_played_questions(self):
        self.count_of_questions_used += 1
        if self.count_of_questions_used == 30:
            winner = sorted(self.list_of_players, key=lambda x: x.score)[-1]
            self.lbl_winner_name.setText(
                f"Победителем стал {winner.name}! Поздравляем!")
            img = QPixmap("application files/win_pic.jpg")
            self.lbl_congratulation.setPixmap(img)
            self.database.add_win(winner)
            for player in self.list_of_players:
                self.database.add_score(player)
            self.stack.setCurrentIndex(4)
            self.clear_gamemaker_data()

    def go_window_with_questions(self):
        self.stack.setCurrentIndex(3)

    def go_game(self):
        self.make_players_in_game()
        self.stack.setCurrentIndex(3)
        self.database.add_players([player.name for player in self.list_of_players])
        self.timer = Timer(self.spinbox_time.value(), self)
        self.timer.start()
        self.timer.count_changed.connect(self.change_progress)

    def go_question(self):
        self.stack.setCurrentIndex(5)

    def add_player(self):
        players = [player.name for player in self.list_of_players]
        if self.lineEdit_player_name.text().strip() != '' and \
                self.lineEdit_player_name.text().strip() not in players:
            self.list_of_players.append(Player(self.lineEdit_player_name.text().title()))
            lbl = QLabel(self.players_names_layout)
            lbl.setText(self.lineEdit_player_name.text())
            self.list_of_lbl_with_players_name.append(lbl)
            self.verticalLayout_5.addWidget(lbl)
            self.number_of_players += 1
            self.lineEdit_player_name.setText("")
            self.list_players_lbl.append(lbl)
            if self.number_of_players == 5:
                self.btn_add_players.setEnabled(False)
        elif self.lineEdit_player_name.text().strip() == '':
            QMessageBox.critical(self, "Ошибка", "Имя не может быть пустым!", QMessageBox.Ok)
        elif self.lineEdit_player_name.text().strip() in players:
            QMessageBox.critical(self, "Ошибка", "Игрок с таким именем уже есть!", QMessageBox.Ok)
        if len(self.list_of_players) == 1:
            self.btn_del_player.setEnabled(True)
            if self.can_start:
                self.btn_start_game.setEnabled(True)
            else:
                self.can_start = True

    def del_player(self):
        self.list_players_lbl[-1].deleteLater()
        self.list_players_lbl.pop(-1)
        self.list_of_players.pop(-1)
        self.number_of_players -= 1

    def closeEvent(self, event):
        self.database.close_database()


def except_hook(cls, exception, traceback):
    sys.__excepthook__(cls, exception, traceback)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    self = MyProgram()
    self.showFullScreen()
    sys.excepthook = except_hook
    sys.exit(app.exec())
